import pandas as pd
import openpyxl
import logging
import re
import os
import tempfile
import shutil
from datetime import datetime
from typing import List, Dict, Any, Union, Optional
from pathlib import Path

class ExcelProcessorV01:
    """
    Enhanced V01 Excel Processor with batch processing and V04 improvements.
    
    Features:
    - Batch processing of multiple Excel files
    - Improved first page generation
    - Enhanced deviation statement handling from V04
    - Better error handling and logging
    """
    
    def __init__(self, output_dir: Optional[str] = None):
        self.logger = logging.getLogger(__name__)
        self.output_dir = output_dir or tempfile.mkdtemp(prefix='bill_processor_')
        os.makedirs(self.output_dir, exist_ok=True)
        
    def process_batch(self, file_paths: List[Union[str, Path]], 
                     output_format: str = 'all') -> Dict[str, Any]:
        """
        Process multiple Excel files in batch mode.
        
        Args:
            file_paths: List of paths to Excel files
            output_format: Output format ('all', 'html', 'pdf', 'docx')
            
        Returns:
            Dict containing processing results and output paths
        """
        results = {
            'success': [],
            'failed': [],
            'output_dir': self.output_dir,
            'start_time': datetime.now().isoformat(),
            'processed_files': 0,
            'total_files': len(file_paths)
        }
        
        for file_path in file_paths:
            try:
                file_path = str(file_path)
                self.logger.info(f"Processing file: {file_path}")
                
                # Process single file
                data = self.process_excel(file_path)
                
                # Generate documents based on format
                output_files = self._generate_documents(data, output_format)
                
                results['success'].append({
                    'file': file_path,
                    'output_files': output_files,
                    'bill_number': data.get('bill_number', 'N/A')
                })
                results['processed_files'] += 1
                
            except Exception as e:
                error_msg = f"Error processing {file_path}: {str(e)}"
                self.logger.error(error_msg, exc_info=True)
                results['failed'].append({
                    'file': file_path,
                    'error': str(e)
                })
        
        results['end_time'] = datetime.now().isoformat()
        results['status'] = 'completed'
        
        # Generate batch summary
        self._generate_batch_summary(results)
        
        return results
    
    def _generate_documents(self, data: Dict[str, Any], 
                          output_format: str = 'all') -> List[str]:
        """
        Generate output documents based on the specified format.
        
        Args:
            data: Processed bill data
            output_format: Output format ('all', 'html', 'pdf', 'docx')
            
        Returns:
            List of generated file paths
        """
        output_files = []
        
        # Create output directory for this bill
        bill_dir = os.path.join(
            self.output_dir, 
            f"bill_{data.get('bill_number', 'unknown')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        )
        os.makedirs(bill_dir, exist_ok=True)
        
        # Generate first page (from V04)
        first_page_path = self._generate_first_page(data, bill_dir)
        if first_page_path:
            output_files.append(first_page_path)
        
        # Generate deviation statement (from V04)
        deviation_path = self._generate_deviation_statement(data, bill_dir)
        if deviation_path:
            output_files.append(deviation_path)
        
        # Generate other document types based on format
        # ... (rest of the document generation logic)
        
        return output_files
    
    def _generate_first_page(self, data: Dict[str, Any], output_dir: str) -> Optional[str]:
        """Generate first page document (V04 style)."""
        try:
            # Create a simple HTML first page
            first_page = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Bill {data.get('bill_number', '')}</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 2cm; }}
                    .header {{ text-align: center; margin-bottom: 2cm; }}
                    .bill-info {{ margin-bottom: 1cm; }}
                    .items-table {{ width: 100%; border-collapse: collapse; margin-top: 1cm; }}
                    .items-table th, .items-table td {{ border: 1px solid #000; padding: 8px; }}
                    .total {{ font-weight: bold; text-align: right; margin-top: 1cm; }}
                </style>
            </head>
            <body>
                <div class="header">
                    <h1>BILL OF QUANTITIES</h1>
                    <h2>{data.get('project_name', '')}</h2>
                </div>
                
                <div class="bill-info">
                    <p><strong>Contractor:</strong> {data.get('contractor_name', '')}</p>
                    <p><strong>Bill Number:</strong> {data.get('bill_number', '')}</p>
                    <p><strong>Date:</strong> {data.get('bill_date', '')}</p>
                </div>
                
                <table class="items-table">
                    <tr>
                        <th>Item No.</th>
                        <th>Description</th>
                        <th>Quantity</th>
                        <th>Unit</th>
                        <th>Rate</th>
                        <th>Amount</th>
                    </tr>
            """
            
            # Add items
            for idx, item in enumerate(data.get('items', [])[:10], 1):  # First 10 items only
                first_page += f"""
                <tr>
                    <td>{idx}</td>
                    <td>{item.get('description', '')}</td>
                    <td>{item.get('quantity', '')}</td>
                    <td>{item.get('unit', '')}</td>
                    <td>{item.get('rate', '')}</td>
                    <td>{item.get('amount', '')}</td>
                </tr>
                """
            
            # Add total
            first_page += f"""
                </table>
                <div class="total">
                    <p>Total Amount: {data.get('total_amount', '0.00')}</p>
                </div>
            </body>
            </html>
            """
            
            # Save HTML file
            output_path = os.path.join(output_dir, 'first_page.html')
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(first_page)
                
            self.logger.info(f"Generated first page: {output_path}")
            return output_path
            
        except Exception as e:
            self.logger.error(f"Error generating first page: {e}")
            return None
    
    def _generate_deviation_statement(self, data: Dict[str, Any], output_dir: str) -> Optional[str]:
        """Generate deviation statement (V04 style)."""
        try:
            # Check if we have deviation items
            deviation_items = data.get('deviation_items', [])
            if not deviation_items:
                self.logger.info("No deviation items found, skipping deviation statement")
                return None
                
            # Create HTML deviation statement
            deviation_doc = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Deviation Statement - Bill {data.get('bill_number', '')}</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 2cm; }}
                    .header {{ text-align: center; margin-bottom: 1cm; }}
                    .deviation-table {{ width: 100%; border-collapse: collapse; margin-top: 1cm; }}
                    .deviation-table th, .deviation-table td {{ 
                        border: 1px solid #000; 
                        padding: 8px;
                        text-align: left;
                    }}
                    .deviation-table th {{ background-color: #f2f2f2; }}
                </style>
            </head>
            <body>
                <div class="header">
                    <h1>DEVIATION STATEMENT</h1>
                    <h2>{data.get('project_name', '')}</h2>
                    <p>Bill No: {data.get('bill_number', '')} | Date: {data.get('bill_date', '')}</p>
                </div>
                
                <table class="deviation-table">
                    <tr>
                        <th>Item No.</th>
                        <th>Description</th>
                        <th>Original Qty</th>
                        <th>Revised Qty</th>
                        <th>Deviation</th>
                        <th>Rate</th>
                        <th>Amount</th>
                    </tr>
            """
            
            # Add deviation items
            for idx, item in enumerate(deviation_items, 1):
                deviation = float(item.get('revised_qty', 0)) - float(item.get('original_qty', 0))
                deviation_doc += f"""
                <tr>
                    <td>{idx}</td>
                    <td>{item.get('description', '')}</td>
                    <td>{item.get('original_qty', '0')}</td>
                    <td>{item.get('revised_qty', '0')}</td>
                    <td>{deviation}</td>
                    <td>{item.get('rate', '0.00')}</td>
                    <td>{item.get('amount', '0.00')}</td>
                </tr>
                """
            
            # Close HTML
            deviation_doc += """
                </table>
            </body>
            </html>
            """
            
            # Save HTML file
            output_path = os.path.join(output_dir, 'deviation_statement.html')
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(deviation_doc)
                
            self.logger.info(f"Generated deviation statement: {output_path}")
            return output_path
            
        except Exception as e:
            self.logger.error(f"Error generating deviation statement: {e}")
            return None
    
    def _generate_batch_summary(self, results: Dict[str, Any]) -> None:
        """Generate a summary of the batch processing."""
        try:
            summary_path = os.path.join(self.output_dir, 'batch_summary.txt')
            with open(summary_path, 'w') as f:
                f.write("=" * 50 + "\n")
                f.write(f"BATCH PROCESSING SUMMARY\n")
                f.write("=" * 50 + "\n\n")
                
                f.write(f"Start Time: {results.get('start_time', 'N/A')}\n")
                f.write(f"End Time: {results.get('end_time', 'N/A')}\n")
                f.write(f"Total Files: {results.get('total_files', 0)}\n")
                f.write(f"Successfully Processed: {len(results.get('success', []))}\n")
                f.write(f"Failed: {len(results.get('failed', []))}\n\n")
                
                if results.get('success'):
                    f.write("\nSUCCESSFULLY PROCESSED FILES:\n")
                    f.write("-" * 50 + "\n")
                    for item in results['success']:
                        f.write(f"- {item['file']} (Bill: {item['bill_number']})\n")
                
                if results.get('failed'):
                    f.write("\nFAILED FILES:\n")
                    f.write("-" * 50 + "\n")
                    for item in results['failed']:
                        f.write(f"- {item['file']}: {item['error']}\n")
            
            results['summary_file'] = summary_path
            self.logger.info(f"Generated batch summary: {summary_path}")
            
        except Exception as e:
            self.logger.error(f"Error generating batch summary: {e}")
    
    def process_excel(self, file_buffer):
        """Process uploaded Excel file and extract all relevant data"""
        try:
            # Load workbook
            workbook = openpyxl.load_workbook(file_buffer, data_only=True)
            
            # Initialize data structure
            data = {
                'project_name': '',
                'contractor_name': '',
                'bill_number': '',
                'bill_date': '',
                'premium_percentage': 0.0,
                'items': [],
                'deviation_items': [],
                'extra_items': [],
                'bill_summary': None,
                'header_data': [],
                'data_frames': {},
                'raw_data': {}
            }
            
            # Process each sheet
            for sheet_name in workbook.sheetnames:
                self.logger.info(f"Processing sheet: {sheet_name}")
                sheet = workbook[sheet_name]
                
                # Convert sheet to DataFrame
                df = self._sheet_to_dataframe(sheet)
                data['data_frames'][sheet_name] = df
                
                # Extract specific data based on sheet content
                self._extract_sheet_data(sheet, df, data, sheet_name)
            
            # Post-process and validate data
            self._post_process_data(data)
            
            workbook.close()
            return data
            
        except Exception as e:
            self.logger.error(f"Error processing Excel file: {e}")
            raise
    
    def _sheet_to_dataframe(self, sheet):
        """Convert Excel sheet to pandas DataFrame"""
        try:
            # Get all values from sheet
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            
            if not data:
                return pd.DataFrame()
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Clean up - remove completely empty rows and columns
            df = df.dropna(how='all').dropna(axis=1, how='all')
            
            return df
            
        except Exception as e:
            self.logger.error(f"Error converting sheet to DataFrame: {e}")
            return pd.DataFrame()
    
    def _extract_sheet_data(self, sheet, df, data, sheet_name):
        """Extract data from specific sheet based on content"""
        
        # Look for header information
        self._extract_header_info(df, data)
        
        # Look for bill items
        if 'bill' in sheet_name.lower() or 'item' in sheet_name.lower():
            items = self._extract_bill_items(df)
            if items:
                data['items'].extend(items)
        
        # Look for deviation items
        if 'deviation' in sheet_name.lower():
            deviation_items = self._extract_deviation_items(df)
            if deviation_items:
                data['deviation_items'].extend(deviation_items)
        
        # Look for extra items
        if 'extra' in sheet_name.lower():
            extra_items = self._extract_extra_items(df)
            if extra_items:
                data['extra_items'].extend(extra_items)
        
        # Look for summary data
        if 'summary' in sheet_name.lower():
            data['bill_summary'] = df
    
    def _extract_header_info(self, df, data):
        """Extract header information from DataFrame"""
        
        # Convert DataFrame to string representation for searching
        df_str = df.astype(str).fillna('')
        
        # Look for project name
        for idx, row in df_str.iterrows():
            for col_idx, cell in enumerate(row):
                if isinstance(cell, str):
                    # Project name patterns
                    if 'project' in cell.lower() and not data['project_name']:
                        next_cell = row.iloc[col_idx + 1] if col_idx + 1 < len(row) else ''
                        if next_cell and next_cell.strip():
                            data['project_name'] = str(next_cell).strip()
                    
                    # Contractor name patterns
                    if 'contractor' in cell.lower() and not data['contractor_name']:
                        next_cell = row.iloc[col_idx + 1] if col_idx + 1 < len(row) else ''
                        if next_cell and next_cell.strip():
                            data['contractor_name'] = str(next_cell).strip()
                    
                    # Bill number patterns
                    if ('bill' in cell.lower() and 'no' in cell.lower()) and not data['bill_number']:
                        next_cell = row.iloc[col_idx + 1] if col_idx + 1 < len(row) else ''
                        if next_cell and next_cell.strip():
                            data['bill_number'] = str(next_cell).strip()
                    
                    # Date patterns
                    if 'date' in cell.lower() and not data['bill_date']:
                        next_cell = row.iloc[col_idx + 1] if col_idx + 1 < len(row) else ''
                        if next_cell and next_cell.strip():
                            data['bill_date'] = str(next_cell).strip()
        
        # Store header data for template rendering
        header_rows = []
        for idx, row in df.iterrows():
            row_data = []
            for cell in row:
                if pd.notna(cell) and str(cell).strip():
                    row_data.append(str(cell).strip())
            if row_data:
                header_rows.append(row_data)
        
        data['header_data'] = header_rows[:10]  # First 10 rows as header
    
    def _extract_bill_items(self, df):
        """Extract bill items from DataFrame"""
        items = []
        
        # Look for columns that might contain item data
        item_cols = self._identify_item_columns(df)
        
        if not item_cols:
            return items
        
        # Extract items
        for idx, row in df.iterrows():
            item = {}
            
            # Basic item structure
            for col_name, col_idx in item_cols.items():
                if col_idx < len(row):
                    value = row.iloc[col_idx]
                    if pd.notna(value):
                        item[col_name] = str(value).strip()
            
            # Only add if we have some data
            if len(item) > 1:  # More than just index
                items.append(item)
        
        return items
    
    def _extract_deviation_items(self, df):
        """Extract deviation items from DataFrame"""
        deviation_items = []
        
        # Look for deviation-specific patterns
        for idx, row in df.iterrows():
            item = {}
            
            # Extract available data
            for col_idx, cell in enumerate(row):
                if pd.notna(cell) and str(cell).strip():
                    if col_idx == 0:
                        item['description'] = str(cell).strip()
                    elif col_idx == 1:
                        item['quantity'] = str(cell).strip()
                    elif col_idx == 2:
                        item['rate'] = str(cell).strip()
                    elif col_idx == 3:
                        item['amount'] = str(cell).strip()
            
            if item:
                deviation_items.append(item)
        
        return deviation_items
    
    def _extract_extra_items(self, df):
        """Extract extra items from DataFrame"""
        extra_items = []
        
        # Similar to deviation items but for extras
        for idx, row in df.iterrows():
            item = {}
            
            # Extract available data
            for col_idx, cell in enumerate(row):
                if pd.notna(cell) and str(cell).strip():
                    if col_idx == 0:
                        item['description'] = str(cell).strip()
                    elif col_idx == 1:
                        item['quantity'] = str(cell).strip()
                    elif col_idx == 2:
                        item['rate'] = str(cell).strip()
                    elif col_idx == 3:
                        item['amount'] = str(cell).strip()
            
            if item:
                extra_items.append(item)
        
        return extra_items
    
    def _identify_item_columns(self, df):
        """Identify columns that contain item data"""
        item_cols = {}
        
        # Look at first few rows to identify headers
        for idx, row in df.head(5).iterrows():
            for col_idx, cell in enumerate(row):
                if pd.notna(cell):
                    cell_str = str(cell).lower().strip()
                    
                    if 'description' in cell_str or 'item' in cell_str:
                        item_cols['description'] = col_idx
                    elif 'quantity' in cell_str or 'qty' in cell_str:
                        item_cols['quantity'] = col_idx
                    elif 'rate' in cell_str:
                        item_cols['rate'] = col_idx
                    elif 'amount' in cell_str:
                        item_cols['amount'] = col_idx
                    elif 'unit' in cell_str:
                        item_cols['unit'] = col_idx
        
        return item_cols
    
    def _post_process_data(self, data):
        """Post-process extracted data"""
        
        # Set defaults if not found
        if not data['project_name']:
            data['project_name'] = 'Project Name Not Found'
        
        if not data['contractor_name']:
            data['contractor_name'] = 'Contractor Name Not Found'
        
        if not data['bill_number']:
            data['bill_number'] = f"BILL-{datetime.now().strftime('%Y%m%d')}"
        
        if not data['bill_date']:
            data['bill_date'] = datetime.now().strftime('%Y-%m-%d')
        
        # Calculate totals
        data['total_amount'] = self._calculate_total_amount(data)
        
        # Add metadata
        data['processing_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        data['processor_version'] = 'V01'
    
    def _calculate_total_amount(self, data):
        """Calculate total amount from all items"""
        total = 0.0
        
        # Sum from regular items
        for item in data['items']:
            amount_str = item.get('amount', '0')
            try:
                amount = float(re.sub(r'[^\d.-]', '', str(amount_str)))
                total += amount
            except:
                pass
        
        # Sum from deviation items
        for item in data['deviation_items']:
            amount_str = item.get('amount', '0')
            try:
                amount = float(re.sub(r'[^\d.-]', '', str(amount_str)))
                total += amount
            except:
                pass
        
        # Sum from extra items
        for item in data['extra_items']:
            amount_str = item.get('amount', '0')
            try:
                amount = float(re.sub(r'[^\d.-]', '', str(amount_str)))
                total += amount
            except:
                pass
        
        return total

# Example usage
if __name__ == "__main__":
    import logging
    logging.basicConfig(level=logging.INFO)
    
    # Example batch processing
    processor = ExcelProcessorV01(output_dir="./output")
    
    # Process a single file
    # result = processor.process_excel("sample_bill.xlsx")
    
    # Or process multiple files in batch
    files = ["bill1.xlsx", "bill2.xlsx"]  # Replace with actual file paths
    results = processor.process_batch(files)
    print(f"Batch processing complete. Results saved in: {results['output_dir']}")
